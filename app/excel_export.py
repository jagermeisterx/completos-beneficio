import io
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from .database import get_donaciones, get_resumen_por_persona, get_total_dinero, get_inventario


HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)
TITLE_FONT = Font(bold=True, size=14, color="2F5496")
SUBTITLE_FONT = Font(bold=True, size=10, color="2F5496")


def _estilo_header(ws, cols):
    for col in range(1, cols + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = BORDER


def _auto_ancho(ws, cols, data_rows):
    for col in range(1, cols + 1):
        max_len = len(str(ws.cell(row=1, column=col).value) or "")
        for row in data_rows:
            val = str(ws.cell(row=row, column=col).value or "")
            max_len = max(max_len, len(val))
        ws.column_dimensions[get_column_letter(col)].width = min(max_len + 4, 50)


def generar_excel():
    wb = Workbook()

    # --- Hoja 1: Donaciones ---
    ws1 = wb.active
    ws1.title = "Donaciones"
    ws1.merge_cells("A1:F1")
    ws1["A1"] = "Registro de Donaciones - Completada a Beneficio"
    ws1["A1"].font = TITLE_FONT

    headers1 = ["ID", "Fecha y Hora", "Nombre", "Apellido", "Producto", "Cantidad"]
    ws1.append(headers1)
    _estilo_header(ws1, len(headers1))

    donaciones = get_donaciones()
    fila = 3
    for d in donaciones:
        ws1.cell(row=fila, column=1, value=d["id"])
        ws1.cell(row=fila, column=2, value=d["fecha_hora"])
        ws1.cell(row=fila, column=3, value=d["nombre"])
        ws1.cell(row=fila, column=4, value=d["apellido"])
        if d["producto_tipo"] == "dinero":
            ws1.cell(row=fila, column=5, value="Dinero ($)")
            ws1.cell(row=fila, column=6, value=f"${d['cantidad']:,}")
        else:
            ws1.cell(row=fila, column=5, value=d["producto_nombre"])
            ws1.cell(row=fila, column=6, value=d["cantidad"])
        for col in range(1, 7):
            ws1.cell(row=fila, column=col).border = BORDER
        fila += 1

    _auto_ancho(ws1, len(headers1), range(3, fila))

    # --- Hoja 2: Resumen por Persona ---
    ws2 = wb.create_sheet("Resumen por Persona")
    ws2.merge_cells("A1:D1")
    ws2["A1"] = "Resumen por Persona - Completada a Beneficio"
    ws2["A1"].font = TITLE_FONT

    headers2 = ["Nombre", "Apellido", "Ítems Donados", "Total a Cobrar ($)"]
    ws2.append(headers2)
    _estilo_header(ws2, len(headers2))

    resumen = get_resumen_por_persona()
    fila2 = 3
    for r in resumen:
        ws2.cell(row=fila2, column=1, value=r["nombre"])
        ws2.cell(row=fila2, column=2, value=r["apellido"])
        ws2.cell(row=fila2, column=3, value=r["items"])
        ws2.cell(row=fila2, column=4, value=f"${r['total_dinero']:,}" if r["total_dinero"] else "$0")
        for col in range(1, 5):
            ws2.cell(row=fila2, column=col).border = BORDER
        fila2 += 1

    # Fila de totales
    total_dinero = get_total_dinero()
    fila2 += 1
    ws2.cell(row=fila2, column=1, value="TOTAL").font = SUBTITLE_FONT
    ws2.cell(row=fila2, column=4, value=f"${total_dinero:,}").font = SUBTITLE_FONT
    for col in range(1, 5):
        ws2.cell(row=fila2, column=col).border = BORDER

    _auto_ancho(ws2, len(headers2), range(3, fila2 + 1))

    # --- Hoja 3: Inventario Final ---
    ws3 = wb.create_sheet("Inventario Final")
    ws3.merge_cells("A1:E1")
    ws3["A1"] = "Inventario Final - Completada a Beneficio"
    ws3["A1"].font = TITLE_FONT

    headers3 = ["Producto", "Requerido", "Donado", "Faltante", "Estado"]
    ws3.append(headers3)
    _estilo_header(ws3, len(headers3))

    inventario = get_inventario()
    fila3 = 3
    for p in inventario:
        if p["tipo"] != "producto":
            continue
        ws3.cell(row=fila3, column=1, value=p["nombre"])
        ws3.cell(row=fila3, column=2, value=p["requerido"])
        ws3.cell(row=fila3, column=3, value=p["donado"])
        ws3.cell(row=fila3, column=4, value=p["faltante"])
        ws3.cell(row=fila3, column=5, value="Completo" if p["completo"] else "Pendiente")
        for col in range(1, 6):
            ws3.cell(row=fila3, column=col).border = BORDER
        fila3 += 1

    _auto_ancho(ws3, len(headers3), range(3, fila3))

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output
